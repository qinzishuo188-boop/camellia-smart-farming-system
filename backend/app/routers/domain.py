import csv
import io
import json
import os
import uuid
from datetime import date, datetime, timedelta
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from sqlmodel import Session, select
from ..algorithms import (
    calculate_confidence,
    calculate_growth_score,
    check_data_completeness,
    compare_growth,
    create_warnings_for_record,
    detect_growth_stage,
    generate_decisions,
    latest_environment,
)
from ..config import get_settings
from ..database import get_session
from ..mode import get_system_mode, is_production, set_system_mode
from ..models import (
    DecisionRecord,
    DiagnosisRecord,
    EnvironmentRecord,
    ExpertReview,
    FarmingLog,
    FileAsset,
    GrowthRecord,
    PestDiseaseKnowledge,
    Plot,
    SensorDevice,
    SystemConfig,
    User,
    WarningRecord,
    WeatherRecord,
)
from ..schemas import (
    DataCompletenessQuery,
    DiagnosisRequest,
    ExecuteDecisionRequest,
    ExpertReviewRequest,
    HandleWarningRequest,
    ImportRowsRequest,
    SystemModeRequest,
    WarningGenerateRequest,
)
from ..security import get_current_user, require_roles

router = APIRouter(tags=["domain"])
settings = get_settings()


def demo_filter(model, session: Session):
    if is_production(session):
        if hasattr(model, "data_source"):
            return select(model).where(model.data_source != "demo")
    return select(model)


@router.get("/dashboard")
def dashboard(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    production = is_production(session)
    base_env_query = select(EnvironmentRecord)
    base_warn_query = select(WarningRecord)
    base_growth_query = select(GrowthRecord)
    base_decision_query = select(DecisionRecord)
    if production:
        base_env_query = base_env_query.where(EnvironmentRecord.data_source != "demo")
        base_warn_query = base_warn_query.where(WarningRecord.data_source != "demo")
        base_growth_query = base_growth_query.where(GrowthRecord.data_source != "demo")
        base_decision_query = base_decision_query.where(DecisionRecord.data_source != "demo")

    plots = session.exec(select(Plot)).all()
    envs = session.exec(base_env_query.order_by(EnvironmentRecord.recorded_at.desc()).limit(500)).all()
    warnings = session.exec(base_warn_query.order_by(WarningRecord.created_at.desc()).limit(200)).all()
    growth = session.exec(base_growth_query.order_by(GrowthRecord.recorded_at.desc()).limit(200)).all()
    decisions = session.exec(base_decision_query.order_by(DecisionRecord.created_at.desc()).limit(20)).all()

    today = datetime.utcnow().date()
    recent_env = [e for e in envs if e.recorded_at.date() >= today - timedelta(days=7)]
    avg = lambda values: round(sum(values) / len(values), 2) if values else 0

    data_available = len(envs) > 0
    if production and not data_available:
        return {
            "mode": "production",
            "data_available": False,
            "guidance": "系统处于生产模式，尚未采集到足够的真实数据。请先创建地块并录入环境记录，或选择进入演示模式查看完整流程。",
            "steps": [
                {"title": "创建油茶地块", "description": "在「地块管理」中录入地块基础档案信息"},
                {"title": "录入环境数据", "description": "在「环境监测」中手动录入或导入环境监测数据"},
                {"title": "记录田间长势", "description": "在「田间巡查」中定期记录新梢、叶片、病虫害等指标"},
                {"title": "生成预警决策", "description": "数据充分后系统将自动生成预警和农事建议"},
            ],
            "cards": {
                "plot_count": len([p for p in plots]),
                "total_area": round(sum(p.area for p in plots), 2),
                "today_monitor_count": 0,
                "current_warning_count": 0,
                "serious_warning_count": 0,
                "today_decision_count": 0,
                "avg_soil_moisture": 0,
                "avg_air_temperature": 0,
                "pest_risk": "未知",
                "growth_score": 0,
            },
            "environment_trend": [],
            "warning_trend": [],
            "risk_levels": {},
            "risk_rank": [],
            "plot_status": [],
            "decisions": [],
            "tasks": [],
        }

    trend = {}
    for e in recent_env:
        key = e.recorded_at.strftime("%m-%d")
        item = trend.setdefault(key, {"date": key, "soil_moisture": [], "air_temperature": [], "air_humidity": []})
        item["soil_moisture"].append(e.soil_moisture)
        item["air_temperature"].append(e.air_temperature)
        item["air_humidity"].append(e.air_humidity)
    trend_rows = [
        {"date": k, "soil_moisture": avg(v["soil_moisture"]), "air_temperature": avg(v["air_temperature"]), "air_humidity": avg(v["air_humidity"])}
        for k, v in sorted(trend.items())
    ]
    risk_rank = []
    for plot in plots:
        plot_warnings = [w for w in warnings if w.plot_id == plot.id and w.status in ["未处理", "处理中"]]
        risk_rank.append({"plot_id": plot.id, "plot_name": plot.plot_name, "count": len(plot_warnings), "serious": sum(1 for w in plot_warnings if w.warning_level == "严重")})
    levels = {name: sum(1 for w in warnings if w.warning_level == name and w.status != "已处理") for name in ["正常", "关注", "警告", "严重"]}
    return {
        "mode": "production" if production else "demo",
        "data_available": True,
        "cards": {
            "plot_count": len(plots),
            "total_area": round(sum(p.area for p in plots), 2),
            "today_monitor_count": sum(1 for e in envs if e.recorded_at.date() == today),
            "current_warning_count": sum(1 for w in warnings if w.status in ["未处理", "处理中"]),
            "serious_warning_count": sum(1 for w in warnings if w.warning_level == "严重" and w.status in ["未处理", "处理中"]),
            "today_decision_count": sum(1 for d in decisions if d.created_at.date() == today),
            "avg_soil_moisture": avg([e.soil_moisture for e in envs[:50]]),
            "avg_air_temperature": avg([e.air_temperature for e in envs[:50]]),
            "pest_risk": "中" if levels.get("警告", 0) else "低",
            "growth_score": avg([g.growth_score for g in growth[:50]]),
        },
        "environment_trend": trend_rows,
        "warning_trend": warning_trend(warnings),
        "risk_levels": levels,
        "risk_rank": sorted(risk_rank, key=lambda x: (x["serious"], x["count"]), reverse=True)[:8],
        "plot_status": risk_rank[:10],
        "decisions": [d.model_dump(mode="json") for d in decisions[:8]],
        "tasks": [w.model_dump(mode="json") for w in warnings if w.status == "未处理"][:8],
    }


def warning_trend(warnings: list[WarningRecord]):
    rows = []
    for i in range(29, -1, -1):
        day = datetime.utcnow().date() - timedelta(days=i)
        rows.append({"date": day.strftime("%m-%d"), "count": sum(1 for w in warnings if w.created_at.date() == day)})
    return rows


@router.get("/system/mode")
def get_mode(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    return {"mode": get_system_mode(session)}


@router.put("/system/mode")
def update_mode(payload: SystemModeRequest, session: Session = Depends(get_session), user: User = Depends(require_roles("admin"))):
    try:
        set_system_mode(session, payload.mode)
        return {"mode": payload.mode, "message": f"系统已切换至{'生产模式' if payload.mode == 'production' else '演示模式'}"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/data-completeness")
def data_completeness(plot_id: int = Query(...), session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    plot = session.get(Plot, plot_id)
    if not plot:
        raise HTTPException(status_code=404, detail="地块不存在")
    return check_data_completeness(session, plot_id)


@router.get("/environment-records")
def list_environment(plot_id: int | None = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    query = demo_filter(EnvironmentRecord, session).order_by(EnvironmentRecord.recorded_at.desc()).limit(500)
    if plot_id:
        query = query.where(EnvironmentRecord.plot_id == plot_id)
    return session.exec(query).all()


@router.post("/environment-records")
def create_environment(payload: EnvironmentRecord, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    session.add(payload)
    session.commit()
    session.refresh(payload)
    created = create_warnings_for_record(session, payload)
    return {"record": payload, "warnings": created}


@router.post("/environment-records/import")
def import_environment(payload: ImportRowsRequest, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    created = []
    errors = []
    for idx, row in enumerate(payload.rows):
        try:
            row.setdefault("data_source", "import")
            item = EnvironmentRecord(**row)
            session.add(item)
            created.append(item)
        except Exception as e:
            errors.append({"row": idx + 1, "error": str(e)})
    session.commit()
    return {"imported": len(created), "errors": errors}


@router.post("/environment-records/import-file")
def import_environment_file(file: UploadFile = File(...), session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    ext = (file.filename or "").split(".")[-1].lower()
    if ext not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 CSV 或 Excel 文件")
    content = file.file.read()
    if ext == "csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row_cells)))
        except ImportError:
            raise HTTPException(status_code=500, detail="服务器缺少 openpyxl 库")
    created = []
    errors = []
    for idx, row in enumerate(rows):
        try:
            row = {k: v for k, v in row.items() if v not in (None, "")}
            if "soil_moisture" in row:
                row["soil_moisture"] = float(row["soil_moisture"])
            if "soil_temperature" in row:
                row["soil_temperature"] = float(row["soil_temperature"])
            if "air_temperature" in row:
                row["air_temperature"] = float(row["air_temperature"])
            if "air_humidity" in row:
                row["air_humidity"] = float(row["air_humidity"])
            if "light_intensity" in row:
                row["light_intensity"] = float(row["light_intensity"])
            if "ph_value" in row:
                row["ph_value"] = float(row["ph_value"])
            if "plot_id" in row:
                row["plot_id"] = int(row["plot_id"])
            row.setdefault("data_source", "import")
            if "recorded_at" in row and row["recorded_at"]:
                row["recorded_at"] = datetime.fromisoformat(row["recorded_at"])
            item = EnvironmentRecord(**row)
            session.add(item)
            created.append(item)
        except Exception as e:
            errors.append({"row": idx + 1, "error": str(e)})
    session.commit()
    return {"imported": len(created), "total": len(rows), "errors": errors}


@router.get("/environment-records/statistics")
def environment_statistics(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    query = demo_filter(EnvironmentRecord, session).order_by(EnvironmentRecord.recorded_at.desc()).limit(300)
    rows = session.exec(query).all()
    avg = lambda values: round(sum(values) / len(values), 2) if values else 0
    return {
        "count": len(rows),
        "soil_moisture": avg([r.soil_moisture for r in rows]),
        "soil_temperature": avg([r.soil_temperature for r in rows]),
        "air_temperature": avg([r.air_temperature for r in rows]),
        "air_humidity": avg([r.air_humidity for r in rows]),
        "light_intensity": avg([r.light_intensity for r in rows]),
        "ph_value": avg([r.ph_value for r in rows]),
    }


@router.get("/environment-records/trend")
def environment_trend(plot_id: int | None = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    rows = list_environment(plot_id, session, user)
    return list(reversed(rows[:60]))


@router.delete("/environment-records/{record_id}")
def delete_environment(record_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    item = session.get(EnvironmentRecord, record_id)
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    session.delete(item)
    session.commit()
    return {"message": "已删除"}


@router.post("/warnings/generate")
def generate_warnings(payload: WarningGenerateRequest, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    records = []
    if payload.plot_id:
        record = latest_environment(session, payload.plot_id)
        records = [record] if record else []
    else:
        query = demo_filter(EnvironmentRecord, session).order_by(EnvironmentRecord.recorded_at.desc()).limit(30)
        records = session.exec(query).all()
    created = []
    for record in records:
        created.extend(create_warnings_for_record(session, record))
    return {"created": len(created), "warnings": created}


@router.get("/warnings")
def list_warnings(level: str | None = None, status: str | None = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    query = demo_filter(WarningRecord, session).order_by(WarningRecord.created_at.desc()).limit(500)
    rows = session.exec(query).all()
    if level:
        rows = [w for w in rows if w.warning_level == level]
    if status:
        rows = [w for w in rows if w.status == status]
    return rows


@router.put("/warnings/{warning_id}/handle")
def handle_warning(warning_id: int, payload: HandleWarningRequest, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    warning = session.get(WarningRecord, warning_id)
    if not warning:
        raise HTTPException(status_code=404, detail="预警不存在")
    warning.status = "已处理"
    warning.handled_by = user.id
    warning.handled_at = datetime.utcnow()
    session.add(warning)
    session.commit()
    return warning


@router.put("/warnings/{warning_id}/ignore")
def ignore_warning(warning_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    warning = session.get(WarningRecord, warning_id)
    if not warning:
        raise HTTPException(status_code=404, detail="预警不存在")
    warning.status = "已忽略"
    session.add(warning)
    session.commit()
    return warning


@router.post("/decision/generate")
def decision_generate(plot_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    plot = session.get(Plot, plot_id)
    if not plot:
        raise HTTPException(status_code=404, detail="地块不存在")
    record = latest_environment(session, plot_id)
    if not record:
        completeness = check_data_completeness(session, plot_id)
        return {
            "data_available": False,
            "completeness": completeness,
            "message": "该地块暂无环境数据，无法生成决策建议。",
            "records": [],
        }
    stage = plot.current_stage or detect_growth_stage(session)
    records = []
    for rule, suggestion_type, content, risk, confidence, rule_version in generate_decisions(session, plot, record):
        item = DecisionRecord(
            plot_id=plot_id,
            stage_name=stage,
            input_json=json.dumps(record.model_dump(mode="json"), ensure_ascii=False),
            matched_rule_id=rule.id if rule else None,
            suggestion_type=suggestion_type,
            suggestion_content=content,
            risk_level=risk,
            confidence=confidence,
            data_source=record.data_source,
            review_status="pending" if rule and rule.requires_expert_review else "pending",
        )
        session.add(item)
        records.append(item)
    session.commit()
    for item in records:
        session.refresh(item)
    completeness = check_data_completeness(session, plot_id)
    return {
        "data_available": True,
        "records": records,
        "completeness": completeness,
    }


@router.get("/decision-records")
def list_decisions(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    query = demo_filter(DecisionRecord, session).order_by(DecisionRecord.created_at.desc()).limit(300)
    return session.exec(query).all()


@router.get("/decision-records/{record_id}")
def get_decision(record_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    item = session.get(DecisionRecord, record_id)
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    return item


@router.put("/decision-records/{record_id}/execute")
def execute_decision(record_id: int, payload: ExecuteDecisionRequest, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    item = session.get(DecisionRecord, record_id)
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    item.is_executed = payload.execution_status in ("executed", "partial")
    item.execution_status = payload.execution_status
    item.execution_cost = payload.execution_cost
    item.execution_image_url = payload.execution_image_url
    item.feedback = payload.feedback
    session.add(item)

    plot = session.get(Plot, item.plot_id)
    log = FarmingLog(
        plot_id=item.plot_id,
        operation_type=item.suggestion_type,
        operation_time=datetime.utcnow(),
        operator=user.username,
        dosage=f"成本 {payload.execution_cost} 元" if payload.execution_cost else None,
        cost=payload.execution_cost,
        description=f"执行系统建议：{item.suggestion_content[:200]}。反馈：{payload.feedback or ''}",
        execution_image_url=payload.execution_image_url,
        data_source="manual",
        linked_decision_id=item.id,
    )
    session.add(log)
    session.commit()
    session.refresh(item)
    return {"decision": item, "farming_log": log}


@router.get("/growth-records")
def list_growth(plot_id: int | None = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    query = demo_filter(GrowthRecord, session).order_by(GrowthRecord.recorded_at.desc()).limit(500)
    if plot_id:
        query = query.where(GrowthRecord.plot_id == plot_id)
    return session.exec(query).all()


@router.post("/growth-records")
def create_growth(payload: GrowthRecord, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    payload.growth_score = calculate_growth_score(payload)
    session.add(payload)
    session.commit()
    session.refresh(payload)
    return payload


@router.put("/growth-records/{record_id}")
def update_growth(record_id: int, payload: dict, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    item = session.get(GrowthRecord, record_id)
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    for key, value in payload.items():
        if hasattr(item, key):
            setattr(item, key, value)
    item.growth_score = calculate_growth_score(item)
    session.add(item)
    session.commit()
    return item


@router.delete("/growth-records/{record_id}")
def delete_growth(record_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    item = session.get(GrowthRecord, record_id)
    if not item:
        raise HTTPException(status_code=404, detail="记录不存在")
    session.delete(item)
    session.commit()
    return {"message": "已删除"}


@router.get("/growth-records/compare")
def growth_compare(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    query = demo_filter(GrowthRecord, session).limit(1000)
    rows = session.exec(query).all()
    result = compare_growth(rows)
    modes = list(result.keys())
    conclusion = "当前数据量不足，建议持续记录后再评估。"
    if "智能管理" in result and "传统管理" in result:
        smart, normal = result["智能管理"], result["传统管理"]
        conclusion = f"本周期内，智能管理区平均新梢长度为 {smart['avg_shoot_length']}cm，传统管理区为 {normal['avg_shoot_length']}cm；智能管理区健康植株率为 {smart['healthy_rate']}%，传统管理区为 {normal['healthy_rate']}%。"
    return {"groups": result, "modes": modes, "conclusion": conclusion}


@router.post("/diagnosis")
def diagnosis(payload: DiagnosisRequest, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    symptoms_text = "、".join(payload.symptoms)
    knowledge = session.exec(select(PestDiseaseKnowledge)).all()
    matched = []
    for item in knowledge:
        if any(s in item.symptoms or s in item.name for s in payload.symptoms):
            matched.append(item)
    if not matched:
        matched = knowledge[:1]
    risk = "严重" if any(s in symptoms_text for s in ["霉变", "枯萎", "树势衰弱"]) else "警告"
    suggestion = "；".join([m.treatment_method for m in matched[:3]]) or "建议补充现场图片并提交农技专家复核。"
    record = DiagnosisRecord(
        plot_id=payload.plot_id,
        symptoms=symptoms_text,
        possible_result="、".join([m.name for m in matched[:3]]) or "待复核",
        risk_level=risk,
        suggestion=suggestion,
        image_url=payload.image_url,
        data_source="manual",
        expert_review="建议复核" if risk == "严重" else "可选复核",
    )
    session.add(record)
    session.commit()
    session.refresh(record)
    return record


@router.get("/diagnosis-records")
def list_diagnosis(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    query = demo_filter(DiagnosisRecord, session).order_by(DiagnosisRecord.created_at.desc()).limit(300)
    return session.exec(query).all()


@router.post("/diagnosis/upload-image")
@router.post("/files/upload")
def upload_file(file: UploadFile = File(...), session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    ext = (file.filename or "").split(".")[-1].lower()
    if ext not in {"jpg", "jpeg", "png", "webp"}:
        raise HTTPException(status_code=400, detail="仅支持 jpg、jpeg、png、webp")
    os.makedirs(settings.upload_dir, exist_ok=True)
    name = f"{uuid.uuid4().hex}.{ext}"
    path = os.path.join(settings.upload_dir, name)
    data = file.file.read()
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件不能超过 5MB")
    with open(path, "wb") as f:
        f.write(data)
    asset = FileAsset(file_name=file.filename or name, file_url=f"/uploads/{name}", file_type=ext, uploaded_by=user.id)
    session.add(asset)
    session.commit()
    session.refresh(asset)
    return asset


@router.get("/files/{file_id}")
def get_file(file_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    item = session.get(FileAsset, file_id)
    if not item:
        raise HTTPException(status_code=404, detail="文件不存在")
    return item


@router.delete("/files/{file_id}")
def delete_file(file_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    item = session.get(FileAsset, file_id)
    if not item:
        raise HTTPException(status_code=404, detail="文件不存在")
    session.delete(item)
    session.commit()
    return {"message": "已删除"}


@router.post("/expert-reviews")
def create_expert_review(payload: ExpertReviewRequest, session: Session = Depends(get_session), user: User = Depends(require_roles("admin", "expert"))):
    review = ExpertReview(
        record_type=payload.record_type,
        record_id=payload.record_id,
        expert_id=user.id,
        review_result=payload.review_result,
        review_comment=payload.review_comment,
    )
    session.add(review)

    if payload.record_type == "decision":
        decision = session.get(DecisionRecord, payload.record_id)
        if decision:
            decision.review_status = payload.review_result
            decision.reviewed_by = user.id
            session.add(decision)
    elif payload.record_type == "diagnosis":
        diagnosis_rec = session.get(DiagnosisRecord, payload.record_id)
        if diagnosis_rec:
            diagnosis_rec.expert_review = payload.review_result
            if payload.review_comment:
                diagnosis_rec.suggestion = f"{diagnosis_rec.suggestion}【专家意见：{payload.review_comment}】"
            session.add(diagnosis_rec)

    session.commit()
    session.refresh(review)
    return review


@router.get("/expert-reviews")
def list_expert_reviews(record_type: str | None = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    query = select(ExpertReview).order_by(ExpertReview.reviewed_at.desc()).limit(300)
    if record_type:
        query = query.where(ExpertReview.record_type == record_type)
    return session.exec(query).all()


@router.get("/expert-reviews/pending")
def pending_reviews(session: Session = Depends(get_session), user: User = Depends(require_roles("admin", "expert"))):
    pending_decisions = session.exec(
        demo_filter(DecisionRecord, session).where(DecisionRecord.review_status == "pending")
    ).all()
    pending_diagnosis = session.exec(
        demo_filter(DiagnosisRecord, session).where(DiagnosisRecord.expert_review.in_(["待复核", "建议复核"]))
    ).all()
    return {
        "pending_decisions": pending_decisions,
        "pending_diagnosis": pending_diagnosis,
    }


@router.get("/farming-logs")
def list_farming_logs(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    query = demo_filter(FarmingLog, session).order_by(FarmingLog.operation_time.desc()).limit(500)
    return session.exec(query).all()


@router.get("/weather-records/latest")
def latest_weather(plot_id: int | None = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    today = date.today()
    query = select(WeatherRecord).where(WeatherRecord.forecast_date >= today).order_by(WeatherRecord.forecast_date)
    if plot_id:
        query = query.where(WeatherRecord.plot_id == plot_id)
    rows = session.exec(query).all()
    result = {}
    for r in rows:
        key = str(r.forecast_date)
        if key not in result:
            result[key] = r
    return list(result.values())


@router.post("/weather/sync")
async def sync_weather(plot_id: int | None = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """Sync weather forecast from QWeather API for plots."""
    from ..weather_service import sync_all_weather, sync_weather_for_plot
    if plot_id:
        plot = session.get(Plot, plot_id)
        if not plot:
            raise HTTPException(status_code=404, detail="地块不存在")
        records = await sync_weather_for_plot(session, plot)
        return {"synced": len(records), "plot_id": plot_id}
    result = await sync_all_weather(session)
    return result


@router.get("/weather/alerts")
async def weather_alerts(plot_id: int, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    """Get weather-based alerts for a plot."""
    from ..weather_service import check_weather_alerts
    plot = session.get(Plot, plot_id)
    if not plot:
        raise HTTPException(status_code=404, detail="地块不存在")
    today = date.today()
    records = session.exec(
        select(WeatherRecord).where(
            WeatherRecord.plot_id == plot_id,
            WeatherRecord.forecast_date >= today,
        ).order_by(WeatherRecord.forecast_date).limit(7)
    ).all()
    alerts = check_weather_alerts(session, plot_id, records)

    from ..models import WarningRecord
    created = []
    for a in alerts:
        warning = WarningRecord(
            plot_id=plot_id,
            warning_type=a["type"],
            warning_level=a["level"],
            warning_content=a["content"],
            suggestion=a["suggestion"],
            trigger_condition=a.get("condition", ""),
            data_source="weather",
        )
        session.add(warning)
        created.append(warning)
    if created:
        session.commit()
        try:
            from ..wecom import send_weather_alert
            await send_weather_alert(
                plot_name=plot.plot_name if plot else f"地块#{plot_id}",
                alert_type=a["type"],
                alert_level=a["level"],
                content=a["content"],
                suggestion=a["suggestion"],
            )
        except Exception:
            pass
    return {"alerts": alerts, "warnings_created": len(created),
            "notifications": {"note": "已生成预警。配置 WxPusher 后可推送微信通知。"}}


@router.get("/notifications/status")
def notification_status(user: User = Depends(get_current_user)):
    """Get notification configuration status."""
    return {
        "wecom_configured": bool(settings.wecom_corp_id and settings.wecom_agent_secret),
        "wecom_corp_id": settings.wecom_corp_id,
        "wecom_agent_id": settings.wecom_agent_id,
        "qrcode_url": "https://work.weixin.qq.com/wework_admin/frame#apps/modApp/1000002",
    }


@router.post("/notifications/test")
async def test_notification(user: User = Depends(get_current_user)):
    """Send a test notification via 企业微信."""
    from ..wecom import send_message
    result = await send_message(
        content=f"""## ✅ 山茶智耘测试通知

**用户：** {user.username}
**时间：** 2026-05-03

这是一条测试消息。

如果收到此消息，说明企业微信通知配置成功！

---
{f"[山茶智耘 · 油茶全周期智慧监测系统]({settings.public_site_url})" if settings.public_site_url else "山茶智耘 · 油茶全周期智慧监测系统"}"""
    )
    return result


@router.get("/reports/export/csv")
@router.get("/reports/export/excel")
def export_csv(session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    query = demo_filter(Plot, session)
    rows = session.exec(query).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["地块编号", "地块名称", "面积", "乡镇", "品种", "管理模式", "当前阶段", "灌溉条件", "土壤pH"])
    for p in rows:
        writer.writerow([p.plot_code, p.plot_name, p.area, p.town, p.variety, p.management_mode, p.current_stage, p.irrigation_type or "-", p.soil_ph or "-"])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=plots.csv"})


@router.get("/reports/export/pdf")
@router.get("/reports/monthly")
@router.get("/reports/yearly")
@router.get("/reports/plot/{plot_id}")
def export_pdf(plot_id: int | None = None, session: Session = Depends(get_session), user: User = Depends(get_current_user)):
    buffer = io.BytesIO()
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    c = canvas.Canvas(buffer, pagesize=A4)
    c.setFont("STSong-Light", 16)
    title = "油茶园智慧监测与管理分析报告"
    if plot_id:
        plot = session.get(Plot, plot_id)
        if plot:
            title = f"{plot.plot_code} {plot.plot_name} 智慧监测与管理分析报告"
    c.drawString(72, 800, title)
    c.setFont("STSong-Light", 11)
    data = dashboard(session, user)
    cards = data.get("cards", {})
    mode = data.get("mode", "demo")
    lines = [
        f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"系统模式：{mode}",
        f"地块总数：{cards.get('plot_count', 0)}，总面积：{cards.get('total_area', 0)} 亩",
        f"当前预警：{cards.get('current_warning_count', 0)}，严重预警：{cards.get('serious_warning_count', 0)}",
        f"平均土壤湿度：{cards.get('avg_soil_moisture', 0)}%，平均气温：{cards.get('avg_air_temperature', 0)}℃",
        f"综合长势评分：{cards.get('growth_score', 0)}",
    ]
    if not data.get("data_available", True):
        lines.append("提示：当前为生产模式，暂无足够真实数据，报告内容有限。")
    lines.append("管理结论：建议持续执行数据采集、预警闭环处理和农事记录，以形成可追溯的油茶全周期管理档案。")
    y = 760
    for line in lines:
        c.drawString(72, y, line)
        y -= 28
    c.showPage()
    c.save()
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=camellia-report.pdf"})


@router.get("/integrations/sensor/upload")
def sensor_placeholder():
    return {"message": "传感器接口已预留，使用 POST /api/environment-records 上传标准监测数据或使用 POST /api/sensor-devices 注册设备。"}


@router.post("/integrations/ai/diagnose")
async def ai_diagnose(payload: dict):
    """Use DeepSeek AI for pest/disease diagnosis assistance."""
    from ..deepseek import diagnose_pest
    symptoms = payload.get("symptoms", [])
    image_desc = payload.get("image_description", "")
    return await diagnose_pest(symptoms, image_desc)


@router.post("/integrations/ai/advice")
async def ai_advice(payload: dict):
    """Use DeepSeek AI for farming advice generation."""
    from ..deepseek import generate_farming_advice
    return {"advice": await generate_farming_advice(
        payload.get("plot_info", {}),
        payload.get("env_data", {}),
        payload.get("growth_data"),
    )}


@router.get("/integrations/ai-image")
def ai_image_placeholder():
    return {
        "message": "AI 图像识别接口已就绪（DeepSeek），使用 POST /api/integrations/ai/diagnose 进行辅助诊断。",
        "provider": "DeepSeek",
        "capabilities": ["病虫害辅助诊断", "农事建议生成", "图片描述分析"],
    }
